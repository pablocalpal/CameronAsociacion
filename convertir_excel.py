import openpyxl
import csv
import os

def main():
    excel_path = 'CameronAsociacion/CAMERON socios.xlsm'
    if not os.path.exists(excel_path):
        excel_path = 'CameronAsociacion/listado socios estructura (1).xlsm'
        
    csv_output_path = 'CameronAsociacion/socios_migracion.csv'
    
    if not os.path.exists(excel_path):
        print(f"Error: No se encontró ningún archivo Excel compatible (CAMERON socios.xlsm o listado socios estructura (1).xlsm)")
        return

    print("Cargando el archivo de Excel...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Columnas exactas de la base de datos para mapeo automático en Supabase
    db_columns = [
        'no',           # A: nº
        'pareja',       # B: Pareja
        'apellido1',    # C: Apellido1
        'apellido2',    # D: Apellido2
        'nombre',       # E: Nombre
        'movil',        # F: Móvil
        'cuota_2024',   # G: CUOTA 2024
        'bajas',        # H: BAJAS
        'no_de_cuenta', # I: Nº de Cuenta
        'direccion',    # J: Dirección
        'cp',           # K: CP
        'poblacion',    # L: Población
        'provincia',    # M: Provincia
        'dni',          # N: DNI
        'regalo',       # O: regalo
        'comida'        # P: comida
    ]

    rows_to_write = []

    # Iterar desde la fila 6 (donde empiezan los datos) hasta el final
    for r in range(6, ws.max_row + 1):
        row_val = []
        has_data = False
        # Leer las columnas A a P (1 a 16)
        for c in range(1, 17):
            val = ws.cell(row=r, column=c).value
            
            # Limpiar float a entero para nº socio y nº pareja si no son nulos
            if c in (1, 2) and isinstance(val, (int, float)):
                val = int(val)
                
            # Limpiar posibles espacios en blanco en strings
            if isinstance(val, str):
                val = val.strip()
                
            if val is not None and val != '':
                has_data = True
                
            row_val.append(val if val is not None else '')
        
        # Solo agregar si la fila tiene algún dato real
        if has_data:
            rows_to_write.append(row_val)

    # Guardar en formato CSV UTF-8 con cabecera
    with open(csv_output_path, mode='w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(db_columns)
        writer.writerows(rows_to_write)

    print(f"¡Hecho! Se han exportado {len(rows_to_write)} socios a '{csv_output_path}'.")

if __name__ == '__main__':
    main()
